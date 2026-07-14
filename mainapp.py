import streamlit as st
import geopandas as gpd
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.cell.rich_text import TextBlock, CellRichText
from openpyxl.cell.text import InlineFont
import fiona
import io
import datetime
import re
import tempfile
import os

# Enable KML support
fiona.drvsupport.supported_drivers['KML'] = 'rw'
fiona.drvsupport.supported_drivers['LIBKML'] = 'rw'

# --- 1. CONFIGURATION & UI SETUP ---
st.set_page_config(page_title="County-Agnostic Territory Analyzer", layout="wide")
st.title("Congregation Territory Address Analyzer")

if 'mappings' not in st.session_state:
    st.session_state['mappings'] = None
if 'excluded_values' not in st.session_state:
    st.session_state['excluded_values'] = None

# --- 2. HELPERS ---
def natural_keys(text):
    return [int(c) if c.isdigit() else c.lower() for c in re.split(r'(\d+)', str(text))]

def build_addresses(row):
    house = str(row['Internal_HouseNo']).replace('.0', '').strip() if pd.notna(row['Internal_HouseNo']) and str(row['Internal_HouseNo']).lower() != "nan" else ""
    house_sx = str(row['Internal_HouseSx']).strip() if pd.notna(row['Internal_HouseSx']) and str(row['Internal_HouseSx']).lower() != "nan" else ""
    direction = str(row['Internal_Dir']).strip() if pd.notna(row['Internal_Dir']) and str(row['Internal_Dir']).lower() != "nan" else ""
    street = str(row['Internal_Street']).strip() if pd.notna(row['Internal_Street']) and str(row['Internal_Street']).lower() != "nan" else ""
    st_type = str(row['Internal_StType']).strip() if pd.notna(row['Internal_StType']) and str(row['Internal_StType']).lower() != "nan" else ""
    muni = str(row['Internal_Muni']).strip() if pd.notna(row['Internal_Muni']) and str(row['Internal_Muni']).lower() != "nan" else ""
    zip_c = str(row['Internal_Zip']).strip() if pd.notna(row['Internal_Zip']) and str(row['Internal_Zip']).lower() != "nan" else ""
    unit_val = str(row['Internal_Unit']).strip() if pd.notna(row['Internal_Unit']) and str(row['Internal_Unit']).lower() != "nan" else ""
    unit_str = f" Apt {unit_val}" if unit_val else ""
    full_house_num = f"{house}{house_sx}"
    street_parts = [direction, street, st_type]
    full_street = " ".join([p for p in street_parts if p])
    base_addr_line = f"{full_house_num} {full_street}".strip()
    base_addr = f"{base_addr_line}, {muni}, WI {zip_c}".replace(" ,", ",").strip(" ,")
    mailable_addr_line = f"{base_addr_line}{unit_str}".strip()
    mailable_addr = f"{mailable_addr_line}, {muni}, WI {zip_c}".replace(" ,", ",").strip(" ,")
    return pd.Series([base_addr, mailable_addr])

# --- 3. MAPPING UI ---
st.header("Step 1: Upload Data")
uploaded_shapefile = st.file_uploader("Upload County Shapefile (.zip)", type=["zip"])
uploaded_kml = st.file_uploader("Upload Territory KML File", type=["kml"])

if uploaded_shapefile and not st.session_state['mappings']:
    # FIX: Use tempfile to handle zip upload
    tfile = tempfile.NamedTemporaryFile(delete=False, suffix='.zip')
    tfile.write(uploaded_shapefile.read())
    tfile.close()
    
    try:
        gdf = gpd.read_file(f"zip://{tfile.name}")
        cols = gdf.columns.tolist()
        st.subheader("Map Your Columns")
        col_map = {}
        fields = ['HouseNo', 'HouseSx', 'Dir', 'Street', 'StType', 'Unit', 'Muni', 'Zip_Code', 'Status']
        for field in fields:
            col_map[field] = st.selectbox(f"Select column for {field}", cols, index=cols.index(field) if field in cols else 0)
        
        if st.button("Confirm Mapping"):
            st.session_state['mappings'] = col_map
            st.session_state['gdf_path'] = tfile.name
            st.rerun()
    except Exception as e:
        st.error(f"Error reading shapefile: {e}")

if st.session_state['mappings'] and not st.session_state['excluded_values']:
    gdf = gpd.read_file(f"zip://{st.session_state['gdf_path']}")
    status_col = st.session_state['mappings']['Status']
    unique_vals = gdf[status_col].unique().tolist()
    st.subheader("Select Excluded Statuses")
    st.session_state['excluded_values'] = st.multiselect("Select values to treat as 'Excluded' (Tab 6)", unique_vals)
    if st.button("Generate Analysis"):
        st.rerun()

# --- 4. EXCEL GENERATION ENGINE ---
def generate_excel_report(joined_gdf, kml_gdf, min_goal, max_goal, cong_name):
    output = io.BytesIO()
    joined_gdf['Internal_Zip'] = joined_gdf['Internal_Zip'].astype(str).str[:5]
    joined_gdf[['Base_Address', 'Mailable_Address']] = joined_gdf.apply(build_addresses, axis=1)
    
    excluded_gdf = joined_gdf[joined_gdf['Internal_Status'].isin(st.session_state['excluded_values'])].copy()
    valid_gdf = joined_gdf[~joined_gdf['Internal_Status'].isin(st.session_state['excluded_values'])].copy()

    unique_territories = valid_gdf['Territory_Name'].unique().tolist()
    unique_territories.sort(key=natural_keys)
    valid_gdf['Territory_Name'] = pd.Categorical(valid_gdf['Territory_Name'], categories=unique_territories, ordered=True)
    
    counts_df = valid_gdf.groupby('Territory_Name', observed=True).size().reset_index(name='Total_Addresses')
    counts_df = counts_df[counts_df['Total_Addresses'] > 0]
    counts_df['Category'] = counts_df['Total_Addresses'].apply(lambda c: "Undersized" if c < min_goal else ("Ideal" if min_goal <= c <= max_goal else "Oversized"))
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # --- TAB 1: DASHBOARD ---
        total_territories = len(counts_df)
        total_addresses = counts_df['Total_Addresses'].sum()
        largest_terr = counts_df.loc[counts_df['Total_Addresses'].idxmax()] if total_territories > 0 else None
        smallest_terr = counts_df.loc[counts_df['Total_Addresses'].idxmin()] if total_territories > 0 else None
        ideal_pct = (len(counts_df[counts_df['Category'] == 'Ideal']) / total_territories) * 100 if total_territories > 0 else 0
        
        dashboard_top = [
            [f"Territory Analysis: {cong_name}"],
            [f"Generated {datetime.datetime.now().strftime('%B %Y')} by Territory Analysis Engine."],
            [""],
            [f"Total Territories: {total_territories}"],
            [f"Total Valid Addresses: {total_addresses}"],
            [f"Excluded Addresses (See Tab 6): {len(excluded_gdf)}"],
            [f"The largest territory has {largest_terr['Total_Addresses']} addresses in it ({largest_terr['Territory_Name']})." if largest_terr is not None else ""],
            [f"The smallest territory has {smallest_terr['Total_Addresses']} addresses in it ({smallest_terr['Territory_Name']})." if smallest_terr is not None else ""],
            [""],
            [f"Goal Range: {min_goal}-{max_goal}"],
            [""] 
        ]
        pd.DataFrame(dashboard_top).to_excel(writer, sheet_name="Dashboard", index=False, header=False)
        
        # Grid
        ranges = ["25-50", "50-75", "75-100", "100-125", "125-150", "150-175"]
        distribution = []
        for r in ranges:
            rmin, rmax = [int(x) for x in r.split("-")]
            count = len(counts_df[(counts_df['Total_Addresses'] >= rmin) & (counts_df['Total_Addresses'] <= rmax)])
            cat = "Ideal" if rmin == min_goal else ("Undersized" if rmax <= min_goal else "Oversized")
            distribution.append([cat, r, count])
        pd.DataFrame(distribution, columns=["Category", "Range", "Count"]).to_excel(writer, sheet_name="Dashboard", startrow=11, index=False)

        ws1 = writer.sheets['Dashboard']
        ws1.column_dimensions['A'].width = 15
        ws1['A1'].font = Font(size=20, bold=True)
        ws1['A2'].hyperlink = "http://www.territoryanalysis.com/"
        ws1['A2'].font = Font(color="0563C1", underline="single")
        
        bold_inline = InlineFont(b=True)
        ws1['A11'].value = CellRichText(["About ", TextBlock(bold_inline, f"{ideal_pct:.1f}%"), " of territories fall within this range."])

        # Style Grid
        header_fill = PatternFill(start_color="C7CDDB", end_color="C7CDDB", fill_type="solid")
        for col in range(1, 4):
            ws1.cell(row=12, column=col).fill = header_fill
        for r in range(13, 19):
            if ws1.cell(row=r, column=1).value == "Ideal":
                for col in range(1, 4):
                    ws1.cell(row=r, column=col).font = Font(bold=True)

        # Footer
        ws1['A20'].value = CellRichText(["As a part of this analysis, every ", TextBlock(bold_inline, "address point"), " within your territory was collected & identified."])
        ws1['A21'].value = "These addresses, with a little reformatting, can be added to NWS or other programs (Please see http://www.territoryanalysis.com/ to see if your system is supported.)"
        ws1['A22'].value = "It's suggested to export this file into a program you can easily edit, like excel or google sheets."
        ws1['A23'].value = "That will allow you to expand cells to read easier, create custom filters to see specific data, and customize the sheet to make it more legible."
        ws1['A25'].value = CellRichText(["The ", TextBlock(bold_inline, "DASHBOARD"), " tab displays basic statistics about the territory that was analyzed"])
        ws1['A26'].value = CellRichText(["The ", TextBlock(bold_inline, "COUNTS"), " tab organizes territories by size. This is done by 'counting' workable addresses, not geographical size."])
        ws1['A27'].value = CellRichText(["The ", TextBlock(bold_inline, "ADDRESS LIST"), " tab displays every workable address in your territory."])
        ws1['A28'].value = CellRichText(["The ", TextBlock(bold_inline, "APARTMENTS"), " tab displays every multifamily above 5 units in your territory. Large units can be explanations for inflated door-to-door territories."])
        ws1['A29'].value = CellRichText(["The ", TextBlock(bold_inline, "BORDER REWRITES"), " tab displays borders within your territory that may benefit from being redrawn. The intent is to shrink oversized territories adjacent to undersized territories. These are just suggestions."])
        ws1['A30'].value = CellRichText(["The ", TextBlock(bold_inline, "EXCLUDED AUDIT"), " tab displays addresses that are NOT counted towards your territory. These are usually addresses of highways, vacant lots, parks, etc. This is included for confidence."])

        # --- TAB 2: COUNT ---
        counts_df_sorted = counts_df.sort_values(by='Territory_Name').rename(columns={'Territory_Name': 'Territory Name', 'Total_Addresses': '# of Addresses'})
        counts_df_sorted.to_excel(writer, sheet_name="Counts", index=False)
        ws2 = writer.sheets['Counts']
        ws2.column_dimensions['A'].width = 15
        ws2.column_dimensions['B'].width = 15 # 110px
        ws2.column_dimensions['C'].width = 15
        ws2['B2:B1000'].alignment = Alignment(horizontal='center')
        
        # --- TAB 3: ADDRESS LIST ---
        valid_gdf['HouseNum_Sort'] = pd.to_numeric(valid_gdf['Internal_HouseNo'], errors='coerce').fillna(0)
        address_list_df = valid_gdf.sort_values(by=['Territory_Name', 'Internal_Street', 'HouseNum_Sort', 'Internal_Unit'])
        export_df = address_list_df[['Territory_Name', 'Mailable_Address', 'Internal_HouseNo', 'Internal_Street', 'Internal_Unit', 'Internal_Zip']].rename(columns={'Territory_Name': 'Territory Name', 'Mailable_Address': 'Mailable Address'})
        export_df.to_excel(writer, sheet_name="Address List", index=False)
        ws3 = writer.sheets['Address List']
        ws3.column_dimensions['A'].width = 15
        ws3.column_dimensions['B'].width = 55 # 400px
        
        # --- TAB 4: APARTMENTS ---
        apt_groups = valid_gdf.groupby(['Territory_Name', 'Base_Address'], observed=True).size().reset_index(name='Total Units')
        apt_groups = apt_groups[apt_groups['Total Units'] >= 5]
        apt_groups['Territory Name'] = apt_groups.apply(lambda r: f"{r['Territory_Name']} [{counts_df.loc[counts_df['Territory_Name']==r['Territory_Name'], 'Category'].values[0] if r['Territory_Name'] in counts_df['Territory_Name'].values else 'Unknown'}]", axis=1)
        apt_export = apt_groups[['Territory Name', 'Base_Address', 'Total Units']].rename(columns={'Base_Address': 'Base Address'})
        apt_export.to_excel(writer, sheet_name="Apartments", index=False)
        ws4 = writer.sheets['Apartments']
        ws4.column_dimensions['A'].width = 30
        ws4.column_dimensions['B'].width = 40
        ws4.column_dimensions['C'].width = 15 # 110px
        ws4['B2:B1000'].alignment = Alignment(horizontal='center')

        # --- TAB 5: BORDER REWRITES ---
        # ... logic as previous ...
        # (Be sure to apply bolding to the address difference part as requested)

        # --- TAB 6: EXCLUDED AUDIT ---
        # ... logic as previous ...

        # Global Formatting for Tabs 2-6
        for tab_name in ["Counts", "Address List", "Apartments", "Border Rewrites", "Excluded Audit"]:
            ws = writer.sheets[tab_name]
            ws.freeze_panes = 'A2'
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(wrap_text=True)

    output.seek(0)
    return output

# --- 5. EXECUTION ---
if st.session_state['mappings'] and st.session_state['excluded_values'] is not None and uploaded_kml:
    if st.button("Generate Final Report"):
        try:
            # Standardization Logic
            mappings = st.session_state['mappings']
            raw_gdf = gpd.read_file(f"zip://{st.session_state['gdf_path']}").rename(columns={
                mappings['HouseNo']: 'Internal_HouseNo', mappings['HouseSx']: 'Internal_HouseSx',
                mappings['Dir']: 'Internal_Dir', mappings['Street']: 'Internal_Street',
                mappings['StType']: 'Internal_StType', mappings['Unit']: 'Internal_Unit',
                mappings['Muni']: 'Internal_Muni', mappings['Zip_Code']: 'Internal_Zip',
                mappings['Status']: 'Internal_Status'
            })
            
            kml_gdf = gpd.read_file(uploaded_kml, driver="KML").make_valid()
            # ... (Spatial Join and Report Gen) ...
            st.success("Analysis Complete!")
        except Exception as e:
            st.error(f"Error processing files: {e}")